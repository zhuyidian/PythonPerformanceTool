from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


class AnalyseResourceExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

        # 创建一个工作簿对象
        wb = Workbook()

        # 调用得到正在运行的工作表。注意：调用工作表的索引默认是0，即默认对第一张工作表进行操作。
        # ws = wb.active
        # ws_1 = wb.create_sheet()  # 默认在结尾处新建一个新的工作表
        # ws_2 = wb.create_sheet(0)  # 在当前工作表的指定索引处新建一个工作表
        # ws_1.title = "新建工作表"  # 用title指定工作表名称
        # 新建工作表，并指定名称
        ws = wb.create_sheet(title=self.sheet_name, index=0)  # 在索引为0的位置创建一个sheet页
        # ws_4 = wb.create_sheet("新建工作表-1", 0)

        # 对sheet页设置一个颜色（16位的RGB颜色）
        ws.sheet_properties.tabColor = 'ff72BA'  # 改变工作表标签颜色，默认为无颜色
        # 将创建的工作簿保存为name.xlsx
        wb.save(f"{self.file_name}.xlsx")
        # 最后关闭文件
        wb.close()

    def open(self):
        name = f"{self.file_name}.xlsx"
        self.wb = load_workbook(filename=name)
        self.sh = self.wb[self.sheet_name]

    def close(self):
        name = f"{self.file_name}.xlsx"
        self.wb.save(name)
        self.wb.close()

    """
    name,cpu_min,cpu_max,cpu_average
    """
    def write_first_row_data(self,data):
        self.first_row_len = len(data)
        self.open()
        for i in range(len(data)):
            self.sh.cell(row=1, column=i + 1, value=data[i])  # 写入数据
        self.close()

    def write_rows_data(self,datas):
        self.open()
        for i in range(len(datas)):
            self.sh.cell(row=2 + i, column=1, value=datas[i].pid_set)
            self.sh.cell(row=2 + i, column=2, value=datas[i].name)
            self.sh.cell(row=2 + i, column=3, value=datas[i].cpu_min)
            self.sh.cell(row=2 + i, column=4, value=datas[i].cpu_max)
            self.sh.cell(row=2 + i, column=5, value=datas[i].cpu_average)
        self.close()

    def write_data(self, row, column, msg):
        self.open()
        self.sh.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)
        self.close()

if __name__ == "__main__":
    analyse_resource_excel = AnalyseResourceExcel('analyse_resource','cpu')
    analyse_resource_excel.write_first_row_data(['name','cpu_min','cpu_max','cpu_average'])